# import os
# from typing import Dict, Optional
# import pandas as pd  # type: ignore
# from dotenv import load_dotenv
# from odf.opendocument import load as load_ods  # type: ignore
# from odf.style import Style, TableCellProperties  # type: ignore
# from odf.table import Table, TableRow, TableCell  # type: ignore
# from odf.text import P  # type: ignore

# load_dotenv()

# _DEFAULT_INPUT = os.getenv("INPUT_EXEL_FILE_PAT")
# _DEFAULT_OUTPUT = os.getenv("OUTPUT_EXEL_FILE_PATH")
# _DEFAULT_SHEET = 0
# _DEFAULT_ROW = 1
# _DEFAULT_MAPPING: Dict[str, str] = {
#     "invoice_no": "Invoice no",
#     "date": "Date",
#     "supplier": "Supplier",
#     "payment_terms": "Payment",
#     "subtotal": "Subtotal",
#     "vat_on_amount": "Vat on Amount",
#     "grand_total_amount": "Grand Total",
# }
# _YELLOW = "#FFFF00"
# _LIGHT_GREEN = "#CCFFCC"


# def _ensure_rows(df: pd.DataFrame, row: int) -> None:
#     while row >= len(df):
#         df.loc[len(df)] = [pd.NA] * len(df.columns)


# def _resolve_sheet_name(source_file: str, sheet_ref: int | str) -> str:
#     """
#     Return the actual sheet name from the workbook, even if caller passed an index.
#     """
#     if isinstance(sheet_ref, str):
#         return sheet_ref
#     with pd.ExcelFile(source_file, engine="odf") as xls:
#         if sheet_ref >= len(xls.sheet_names) or sheet_ref < 0:
#             raise IndexError(f"Sheet index {sheet_ref} out of range.")
#         return xls.sheet_names[sheet_ref]


# def _add_new_sheet(
#     ods_path: str,
#     payload: Dict[str, str],
#     mapping: Dict[str, str],
#     source_sheet_name: str,
# ) -> None:
#     doc = load_ods(ods_path)

#     header_style = Style(name="HeaderYellow", family="table-cell")
#     header_style.addElement(TableCellProperties(backgroundcolor=_YELLOW))
#     data_style = Style(name="RowGreen", family="table-cell")
#     data_style.addElement(TableCellProperties(backgroundcolor=_LIGHT_GREEN))
#     doc.styles.addElement(header_style)
#     doc.styles.addElement(data_style)
#     for tbl in list(doc.spreadsheet.childNodes):
#         if tbl.qname[1] == "table" and tbl.getAttribute("name") == "week_1":
#             doc.spreadsheet.removeChild(tbl)

#     summary = Table(name="week_1")
#     doc.spreadsheet.addElement(summary)

#     head_row = TableRow()
#     for col_name in mapping.values():
#         cell = TableCell(stylename=header_style)
#         cell.addElement(P(text=str(col_name)))
#         head_row.addElement(cell)
#     summary.addElement(head_row)

#     data_row = TableRow()
#     for key in mapping.keys():
#         cell = TableCell(stylename=data_style)
#         cell.addElement(P(text=str(payload.get(key, ""))))
#         data_row.addElement(cell)
#     summary.addElement(data_row)

#     doc.save(ods_path)


# def write_agents_data_into_excel(
#     payload: Dict[str, str],
#     *,
#     input_path: str = _DEFAULT_INPUT,
#     output_path: str = _DEFAULT_OUTPUT,
#     sheet: int | str = _DEFAULT_SHEET,
#     row: int = _DEFAULT_ROW,
#     mapping: Optional[Dict[str, str]] = None,
#     add_summary: bool = True,
# ) -> str:
#     mapping = mapping or _DEFAULT_MAPPING

#     df = pd.read_excel(input_path, sheet_name=sheet, engine="odf")
#     _ensure_rows(df, row)

#     for key, column in mapping.items():
#         if key not in payload:
#             raise KeyError(f"payload missing key '{key}'")
#         if column not in df.columns:
#             raise KeyError(f"spreadsheet missing column '{column}'")
#         df.at[row, column] = payload[key]

#     os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
#     df.to_excel(output_path, sheet_name=sheet, index=False, engine="odf")

#     if add_summary:
#         real_name = _resolve_sheet_name(input_path, sheet)
#         _add_new_sheet(output_path, payload, mapping, real_name)

#     return os.path.abspath(output_path)





# import os
# from typing import Dict, Optional
# import pandas as pd
# from dotenv import load_dotenv
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill, Alignment, Font

# load_dotenv()

# _DEFAULT_INPUT  = os.getenv("INPUT_EXCEL_FILE_PATH")
# _DEFAULT_OUTPUT = os.getenv("OUTPUT_EXCEL_FILE_PATH")
# _DEFAULT_SHEET  = 0
# _DEFAULT_ROW    = 1
# _DEFAULT_MAPPING: Dict[str,str] = {
#     "invoice_no":       "Invoice no",
#     "date":             "Date",
#     "supplier":         "Supplier",
#     "payment_terms":    "Payment",
#     "subtotal":         "Subtotal",
#     "vat_on_amount":    "Vat on Amount",
#     "grand_total_amount":"Grand Total",
# }

# _YELLOW      = PatternFill("solid", fgColor="FFFF00")
# _LIGHT_GREEN = PatternFill("solid", fgColor="CCFFCC")
# _CENTER      = Alignment(horizontal="center", vertical="center")

# def _ensure_rows(df: pd.DataFrame, row: int) -> None:
#     while row >= len(df):
#         df.loc[len(df)] = [pd.NA]*len(df.columns)

# def write_agents_data_into_excel_xlsx(
#     payload: Dict[str,str],
#     *,
#     input_path:  str            = _DEFAULT_INPUT,
#     output_path: str            = _DEFAULT_OUTPUT,
#     sheet:       int|str        = _DEFAULT_SHEET,
#     row:         int            = _DEFAULT_ROW,
#     mapping:     Optional[Dict[str,str]] = None,
#     add_summary: bool           = True,
# ) -> str:
#     mapping = mapping or _DEFAULT_MAPPING

#     # 1) load into pandas, ensure enough rows
#     df = pd.read_excel(input_path, sheet_name=sheet)
#     _ensure_rows(df, row)

#     # 2) write payload into row
#     for key, col in mapping.items():
#         if key not in payload:
#             raise KeyError(f"payload missing key '{key}'")
#         if col not in df.columns:
#             raise KeyError(f"spreadsheet missing column '{col}'")
#         df.at[row, col] = payload[key]

#     # 3) save to XLSX via openpyxl engine
#     os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
#     with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
#         df.to_excel(writer, sheet_name=str(sheet), index=False)

#     # 4) optionally append a summary sheet with styles
#     if add_summary:
#         wb = load_workbook(output_path)
#         summary_name = "Summary"
#         if summary_name in wb.sheetnames:
#             del wb[summary_name]
#         ws = wb.create_sheet(summary_name)

#         # header row
#         for c_idx, header in enumerate(mapping.values(), start=1):
#             cell = ws.cell(row=1, column=c_idx, value=header)
#             cell.fill    = _YELLOW
#             cell.alignment= _CENTER
#             cell.font     = Font(bold=True)

#         # data row
#         for c_idx, key in enumerate(mapping.keys(), start=1):
#             cell = ws.cell(row=2, column=c_idx, value=payload.get(key, ""))
#             cell.fill    = _LIGHT_GREEN
#             cell.alignment= _CENTER

#         # auto-fit column widths
#         for col in ws.columns:
#             max_len = max(len(str(cell.value or "")) for cell in col)
#             ws.column_dimensions[col[0].column_letter].width = max_len + 2

#         wb.save(output_path)

#     return os.path.abspath(output_path)



import os
from typing import Dict, Optional
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

load_dotenv()

_DEFAULT_INPUT  = os.getenv("INPUT_EXEL_FILE_PAT")
_DEFAULT_OUTPUT = os.getenv("OUTPUT_EXEL_FILE_PATH")
_DEFAULT_SHEET  = 0
_DEFAULT_ROW    = 1
_DEFAULT_MAPPING: Dict[str,str] = {
    "invoice_no":       "Invoice no",
    "date":             "Date",
    "supplier":         "Supplier",
    "payment_terms":    "Payment",
    "subtotal":         "Subtotal",
    "vat_on_amount":    "Vat on Amount",
    "grand_total_amount":"Grand Total",
}

_YELLOW      = PatternFill("solid", fgColor="FFFF00")
_LIGHT_GREEN = PatternFill("solid", fgColor="CCFFCC")
_CENTER      = Alignment(horizontal="center", vertical="center")

def _ensure_rows(df: pd.DataFrame, row: int) -> None:
    while row >= len(df):
        df.loc[len(df)] = [pd.NA]*len(df.columns)

def write_agents_data_into_excel(
    payload: Dict[str,str],
    *,
    input_path:  str            = _DEFAULT_INPUT,
    output_path: str            = _DEFAULT_OUTPUT,
    sheet:       int|str        = _DEFAULT_SHEET,
    row:         int            = _DEFAULT_ROW,
    mapping:     Optional[Dict[str,str]] = None,
    add_summary: bool           = True,
) -> str:
    mapping = mapping or _DEFAULT_MAPPING

    df = pd.read_excel(input_path, sheet_name=sheet)
    _ensure_rows(df, row)
    for key, col in mapping.items():
        if key not in payload:
            raise KeyError(f"payload missing key '{key}'")
        if col not in df.columns:
            raise KeyError(f"spreadsheet missing column '{col}'")
        df.at[row, col] = payload[key]

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, sheet_name=str(sheet), index=False)

    if add_summary:
        wb = load_workbook(output_path)
        summary_name = "week_1"
        if summary_name in wb.sheetnames:
            del wb[summary_name]
        ws = wb.create_sheet(summary_name)
        for c_idx, header in enumerate(mapping.values(), start=1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.fill    = _YELLOW
            cell.alignment= _CENTER
            cell.font     = Font(bold=True)

        for c_idx, key in enumerate(mapping.keys(), start=1):
            cell = ws.cell(row=2, column=c_idx, value=payload.get(key, ""))
            cell.fill    = _LIGHT_GREEN
            cell.alignment= _CENTER
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        wb.save(output_path)

    return os.path.abspath(output_path)
